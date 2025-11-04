import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE_00

def create_excel_report(df_summary, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # --- Заголовок ---
    headers = df_summary.columns.tolist()
    ws.append(headers)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --- Дані ---
    df_display = df_summary.copy()

    # Перетворюємо числові колонки у float для форматування
    numeric_cols = ['unit_price', 'cost_per_unit', 'total_sales', 'total_profit', 'avg_profit_margin']
    for col in numeric_cols:
        df_display[col] = pd.to_numeric(df_display[col].replace(r'[\$,]', '', regex=True), errors='coerce')

    # Додаємо дані до Excel
    for row in df_display.values.tolist():
        ws.append(row)

    # --- Форматування рядків ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for idx, cell in enumerate(row):
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="2C2C2C", end_color="2C2C2C", fill_type="solid")

            # Форматування чисел
            col_name = headers[idx]
            if col_name in ['unit_price', 'cost_per_unit', 'total_sales', 'total_profit']:
                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                cell.alignment = Alignment(horizontal="right")
            elif col_name == 'avg_profit_margin':
                cell.number_format = "0.0\"%\""
                cell.alignment = Alignment(horizontal="right")

    # --- Автоширина колонок ---
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # +2 для відступу

    # Зберігаємо файл
    wb.save(excel_path)

